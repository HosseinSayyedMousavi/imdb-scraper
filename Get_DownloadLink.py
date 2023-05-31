import requests
from bs4 import BeautifulSoup
import json
from pprintpp import pprint
import sqlite3
import openpyxl
headers = {
  'authority': 'www.imdb.com',
  'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
  'accept-language': 'en-US,en;q=0.9,fa;q=0.8',
  'sec-ch-ua': '"Google Chrome";v="113", "Chromium";v="113", "Not-A.Brand";v="24"',
  'sec-ch-ua-mobile': '?0',
  'sec-ch-ua-platform': '"Windows"',
  'sec-fetch-dest': 'document',
  'sec-fetch-mode': 'navigate',
  'sec-fetch-site': 'same-origin',
  'sec-fetch-user': '?1',
  'upgrade-insecure-requests': '1',
  'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36',
  'Cookie': 'session-id=139-2944897-7912339; session-id-time=2082787201l'
}

workbook = openpyxl.load_workbook("download_list.xlsx")
sheet = workbook.active
n = sheet.max_row
for i in range(2,n+1):
    if sheet.cell(row=i , column = 1).value.strip():
        url = sheet.cell(row=i , column = 1).value
        print(i-1 , url)
        try:
            r = requests.request("GET", url, headers=headers)
            soup = BeautifulSoup(r.text,"html.parser")
            element = json.loads(soup.select("script[type='application/ld+json']")[0].text)
            url2 =element["trailer"]["url"]
            r2 = requests.request("GET", url2, headers=headers)
            soup = BeautifulSoup(r2.text,"html.parser")
            element = json.loads(soup.select("script[id='__NEXT_DATA__']")[0].text)
            DownloadLink=element["props"]["pageProps"]["videoPlaybackData"]["video"]["playbackURLs"][1]["url"]
            sheet.cell(row=i , column = 2).value = DownloadLink
            workbook.save("download_list.xlsx")
            print("\n"+DownloadLink+"\n")
        except:
            print("\nNOT EXISTS OR NOT FOUND!\n")